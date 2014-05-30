# Create an Excel spreadsheet containing 96 construct variants of human Abl1 (P00519)
#
# Daniel L. Parton <partond@mskcc.org> - 5 Jul 2013
#

import sys,os
from openpyxl import Workbook
from lxml import etree
import choderalab as clab
from numpy import *
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
from random import sample

spreadsheet_path = 'abl1-constructs.xlsx'

# =============
# kinDB_id of pk_domain for which to generate construct variants
# NOTE: certain parts of the script are specific to ABL1_HUMAN, so this script should not be used unmodified for other kinases
# =============
query_kinase_id = 'ABL1_HUMAN_PK0_P00519'
print 'Generating construct variants for protein:', query_kinase_id

# =============
# Parse kinDB
# =============
kinDBpath = os.path.join(clab.kinome_rootdir, 'database', 'kinDB-complete.xml')
kinDB = etree.parse(kinDBpath).getroot()

# =============
# Get the UniProt resnums of the start and end of the pk_domain
# =============
query_kinase = kinDB.find('kinase/uniprot/pk_domain[@kinDB_id="%s"]' % query_kinase_id)
uniprot_start = int( query_kinase.get('begin') )
uniprot_end = int( query_kinase.get('end') )
print 'Uniprot start and end:', uniprot_start, uniprot_end

# =============
# Generate 81 (9x9) construct variants
# =============

seq_starts = zeros(9, dtype='int64')
seq_starts[0] = uniprot_start - 14
seq_starts[1] = uniprot_start - 13 # start of most PDB constructs
seq_starts[2] = uniprot_start - 12
seq_starts[3] = uniprot_start - 5
seq_starts[4] = uniprot_start - 4
seq_starts[5] = uniprot_start - 2
seq_starts[6] = uniprot_start - 1
seq_starts[7] = uniprot_start
seq_starts[8] = uniprot_start + 1

seq_ends = zeros(9, dtype='int64')
seq_ends[0] = uniprot_end - 3
seq_ends[1] = uniprot_end
seq_ends[2] = uniprot_end + 1
seq_ends[3] = uniprot_end + 3
seq_ends[4] = uniprot_end + 7 # end of 3CS9 construct
seq_ends[5] = uniprot_end + 18 # end of C-terminal helix according to 2FO0
seq_ends[6] = uniprot_end + 19 # end of 2G2H construct
seq_ends[7] = uniprot_end + 20
seq_ends[8] = uniprot_end + 22 # end of 2E2B construct

# =============
# Choose 3 constructs from PDB structures to use as control sequences (i.e. sequences which are known to express)
# Create 5 replicas for each construct
# NOTE: manual intervention is used to choose these constructs. This script will not work for other kinases without modification.
# =============

control_starts = [229,229,229]
control_ends = [500,512,515]

# =============
# Distribute the 96 constructs in a flattened array, and in plate format
# =============

# Plate ordering:
# One of each of the three control constructs goes at the beginning of the plate
# One of each of the three control constructs goes at the end of the plate

plate_seq_starts_flat = zeros(96, dtype='int64')
plate_seq_starts_flat[0] = control_starts[0]
plate_seq_starts_flat[1] = control_starts[1]
plate_seq_starts_flat[2] = control_starts[2]
plate_seq_starts_flat[-1] = control_starts[2]
plate_seq_starts_flat[-2] = control_starts[1]
plate_seq_starts_flat[-3] = control_starts[0]

plate_seq_ends_flat = zeros(96, dtype='int64')
plate_seq_ends_flat[0] = control_ends[0]
plate_seq_ends_flat[1] = control_ends[1]
plate_seq_ends_flat[2] = control_ends[2]
plate_seq_ends_flat[-1] = control_ends[2]
plate_seq_ends_flat[-2] = control_ends[1]
plate_seq_ends_flat[-3] = control_ends[0]

# A further three of each of the control constructs are distributed randomly within the rest of the plate
randints = sample(range(3,93),9)
for i in range(0,9,3):
    plate_seq_starts_flat[randints[i]] = control_starts[0]
    plate_seq_starts_flat[randints[i+1]] = control_starts[1]
    plate_seq_starts_flat[randints[i+2]] = control_starts[2]
    plate_seq_ends_flat[randints[i]] = control_ends[0]
    plate_seq_ends_flat[randints[i+1]] = control_ends[1]
    plate_seq_ends_flat[randints[i+2]] = control_ends[2]

# The manually-chosen construct variants are then distributed according to the ordering above, excepting insertions of the control constructs
seq_starts_81 = zeros(81, dtype='int64')
for i in range(len(seq_starts)):
    for j in range(len(seq_ends)):
        seq_starts_81[ (i * len(seq_starts)) + j ] = seq_starts[i]

i = 3
x = 0
while x < 81:
    while plate_seq_starts_flat[i] != 0:
        i += 1
    plate_seq_starts_flat[i] = seq_starts_81[x]
    x += 1

seq_ends_81 = zeros(81, dtype='int64')
for i in range(len(seq_ends)):
    for j in range(len(seq_ends)):
        seq_ends_81[ (i * len(seq_ends)) + j ] = seq_ends[j]

i = 3
x = 0
while x < 81:
    while plate_seq_ends_flat[i] != 0:
        i += 1
    plate_seq_ends_flat[i] = seq_ends_81[x]
    x += 1

#plate_seq_starts_flat = zeros(96, dtype='int64')
#for i in range(len(seq_starts)):
#    for j in range(len(seq_ends)):
#        plate_seq_starts_flat[ (i * len(seq_starts)) + j ] = seq_starts[i]
#
#for i in range(3):
#    for j in range(5):
#        plate_seq_starts_flat[(len(seq_starts) ** 2) + ((i * 5) + j)] = control_starts[i]
#
#plate_seq_ends_flat = zeros(96, dtype='int64')
#for i in range(len(seq_starts)):
#    for j in range(len(seq_ends)):
#        plate_seq_ends_flat[ (i * len(seq_starts)) + j ] = seq_ends[j]
#
#for i in range(3):
#    for j in range(5):
#        plate_seq_ends_flat[(len(seq_ends) ** 2) + ((i * 5) + j)] = control_ends[i]

# NOTE: the plate format array is not actually used here, since the MacroLab take a flattened list of constructs.

plate_seq_starts = reshape(plate_seq_starts_flat, (8,12))
plate_seq_ends = reshape(plate_seq_ends_flat, (8,12))
print plate_seq_starts
print plate_seq_ends

# =============
# Get the QB3 macrolab Abl1 DNA sequence, convert to aa sequence, and map the residue positions
# =============

uniprot_seq = clab.core.sequnwrap(query_kinase.find('../sequence').text)

with open('abl1-macrolab-dna_seq.txt','r') as macrolab_dna_file:
    macrolab_dna_seq = macrolab_dna_file.readline().strip()

translated_macrolab_dna = Seq(macrolab_dna_seq, IUPAC.unambiguous_dna).translate()

alignment = clab.align.run_clustalo(['A','B'], [uniprot_seq, translated_macrolab_dna])

print alignment[0]
print alignment[1]
print clab.align.seq_comparator(alignment[0], alignment[1])

# NOTE: protein sequences are identical (except for additional C-term Leu in macrolab sequence), so have not written any code for mapping the residue positions

# =============
# Make Excel spreadsheet
# =============

wb = Workbook()
ws = wb.get_active_sheet()
ws.title = 'Abl1_constructs'

# Also print data to a csv file
csv_file = open('abl1-constructs-data.csv','w')
csv_file.write('Construct index, construct start residue (1-based aa), construct end residue, construct aa sequence, construct DNA sequence (from Harvard DF/HCC library of pJP1520 kinase plasmids)\n')

for i in range(96):
    title_cell = ws.cell(row=i, column=0)
    title_cell.value = '%d-%d' % (plate_seq_starts_flat[i], plate_seq_ends_flat[i])
    seq_cell = ws.cell(row=i, column=1)
    construct_aa_seq = uniprot_seq[ plate_seq_starts_flat[i] - 1 : plate_seq_ends_flat[i] ] # DEBUG
    construct_dna_seq = macrolab_dna_seq[ ( plate_seq_starts_flat[i] - 1 ) * 3 : plate_seq_ends_flat[i] * 3 ]
    translated_construct_dna_seq = Seq(construct_dna_seq, IUPAC.unambiguous_dna).translate() # DEBUG
    #print clab.align.seq_comparator(construct_aa_seq, translated_construct_dna_seq) # DEBUG
    #print plate_seq_ends_flat[i] - plate_seq_starts_flat[i] + 1, len(construct_dna_seq)/3, len(construct_dna_seq) % 3 # DEBUG
    seq_cell.value = construct_dna_seq
    csv_file.write('%d,%d,%d,%s,%s\n' % (i, plate_seq_starts_flat[i], plate_seq_ends_flat[i], construct_aa_seq, construct_dna_seq))



# NOTE: this makes an entirely randomized plate
#random_ints = sample(range(96),96)
#for i in range(96):
#    title_cell = ws.cell(row=random_ints[i], column=0)
#    title_cell.value = '%d-%d' % (plate_seq_starts_flat[i], plate_seq_ends_flat[i])
#    seq_cell = ws.cell(row=random_ints[i], column=1)
#    construct_aa_seq = uniprot_seq[ plate_seq_starts_flat[i] - 1 : plate_seq_ends_flat[i] ] # DEBUG
#    construct_dna_seq = macrolab_dna_seq[ ( plate_seq_starts_flat[i] - 1 ) * 3 : plate_seq_ends_flat[i] * 3 ]
#    translated_construct_dna_seq = Seq(construct_dna_seq, IUPAC.unambiguous_dna).translate() # DEBUG
#    #print clab.align.seq_comparator(construct_aa_seq, translated_construct_dna_seq) # DEBUG
#    #print plate_seq_ends_flat[i] - plate_seq_starts_flat[i] + 1, len(construct_dna_seq)/3, len(construct_dna_seq) % 3 # DEBUG
#    seq_cell.value = construct_dna_seq
#    csv_file.write('%d,%d,%d,%d,%s,%s\n' % (i, random_ints[i], plate_seq_starts_flat[i], plate_seq_ends_flat[i], construct_aa_seq, construct_dna_seq))
#
csv_file.close()

# Save spreadsheet
wb.save(spreadsheet_path)


