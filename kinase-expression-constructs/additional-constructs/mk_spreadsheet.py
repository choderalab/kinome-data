import sys, os, re
from openpyxl import Workbook
import pandas as pd

def main():
    output_xl_filepath = 'collab-kinases.xlsx'

    addgene_plasmids_df = pd.read_csv('../../plasmids/addgene/human-kinase-ORF-collection/aln.csv')
    hip_plasmids_df = pd.read_csv('../../plasmids/DFHCC-PlasmID/HIP-human_kinase_collection-pJP1520/aln.csv')

    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = 'kinase constructs'

    targets = [
        {
            'targetid': 'BTK_HUMAN_D0',
            'plasmid_source': 'addgene',
            'cloneID': '23918',
            'construct_plasmid_aa_seq': 'IDPKDLTFLKELGTGQFGVVKYGKWRGQYDVAIKMIKEGSMSEDEFIEEAKVMMNLSHEKLVQLYGVCTKQRPIFIITEYMANGCLLNYLREMRHRFQTQQLLEMCKDVCEAMEYLESKQFLHRDLAARNCLVNDQGVVKVSDFGLSRYVLDDEYTSSVGSKFPVRWSPPEVLMYSKFSSKSDIWAFGVLMWEIYSLGKMPYERFTNSETAEHIAQGLRLYRPHLASEKVYTIMYSCWHEKADERPTFKILLSNILDVMDEES',
        },
        {
            'targetid': 'PDPK1_HUMAN_D0',
            'plasmid_source': 'HIP',
            'cloneID': 'HsCD00037925',
            'construct_plasmid_aa_seq': 'GPAMDGTAAEPRPGAGSLQHAQPPPQPRKKRPEDFKFGKILGEGSFSTVVLARELATSREYAIKILEKRHIIKENKVPYVTRERDVMSRLDHPFFVKLYFTFQDDEKLYFGLSYAKNGELLKYIRKIGSFDETCTRFYTAEIVSALEYLHGKGIIHRDLKPENILLNEDMHIQITDFGTAKVLSPESKQARANSFVGTAQYVSPELLTEKSACKSSDLWALGCIIYQLVAGLPPFRAGNEYLIFQKIIKLEYDFPEKFFPKARDLVEKLLVLDATKRLGCEEMEGYGPLKAHPFFESVTWENLHQQTPPKLT',
        },
        {
            'targetid': 'ERBB2_HUMAN_D0',
            'plasmid_source': 'addgene',
            'cloneID': '23888',
            'construct_plasmid_aa_seq': 'MELAALCRWGLLLALLPPGAASTQVCTGTDMKLRLPASPETHLDMLRHLYQGCQVVQGNLELTYLPTNASLSFLQDIQEVQGYVLIAHNQVRQVPLQRLRIVRGTQLFEDNYALAVLDNGDPLNNTTPVTGASPGGLRELQLRSLTEILKGGVLIQRNPQLCYQDTILWKDIFHKNNQLALTLIDTNRSRACHPCSPMCKGSRCWGESSEDCQSLTRTVCAGGCARCKGPLPTDCCHEQCAAGCTGPKHSDCLACLHFNHSGICELHCPALVTYNTDTFESMPNPEGRYTFGASCVTACPYNYLSTDVGSCTLVCPLHNQEVTAEDGTQRCEKCSKPCARVCYGLGMEHLREVRAVTSANIQEFAGCKKIFGSLAFLPESFDGDPASNTAPLQPEQLQVFETLEEITGYLYISAWPDSLPDLSVFQNLQVIRGRILHNGAYSLTLQGLGISWLGLRSLRELGSGLALIHHNTHLCFVHTVPWDQLFRNPHQALLHTANRPEDECVGEGLACHQLCARGHCWGPGPTQCVNCSQFLRGQECVEECRVLQGLPREYVNARHCLPCHPECQPQNGSVTCFGPEADQCVACAHYKDPPFCVARCPSGVKPDLSYMPIWKFPDEEGACQPCPINCTHSCVDLDDKGCPAEQRASPLTSIISAVVGILLVVVLGVVFGILIKRRQQKIRKYTMRRLLQETELVEPLTPSGAMPNQAQMRILKETELRKVKVLGSGAFGTVYKGIWIPDGENVKIPVAIKVLRENTSPKANKEILDEAYVMAGVGSPYVSRLLGICLTSTVQLVTQLMPYGCLLDHVRENRGRLGSQDLLNWCMQIAKGMSYLEDVRLVHRDLAARNVLVKSPNHVKITDFGLARLLDIDETEYHADGGKVPIKWMALESILRRRFTHQSDVWSYGVTVWELMTFGAKPYDGIPAREIPDLLEKGERLPQPPICTIDVYMIMVKCWMIDSECRPRFRELVSEFSRMARDPQRFVVIQNEDLGPASPLDSTFYRSLLEDDDMGDLVDAEEYLVPQQGFFCPDPAPGAGGMVHHRHRSSSTRSGGGDLTLGLEPSEEEAPRSPLAPSEGAGSDVFDGDLGMGAAKGLQSLPTHDPSPLQRYSEDPTVPLPSETDGYVAPLTCSPQPEYVNQPDVRPQPPSPREGPLPAARPAGATLERPKTLSPGKNGVVKDVFAFGGAVENPEYLTPQGGAAPQPHPPPAFSPAFDNLYYWDQDPPERGAPPSTFKGTPTAENPEYLGLDVPV',
        },
        {
            'targetid': 'MP2K1_HUMAN_D0',
            'plasmid_source': 'addgene',
            'cloneID': '23406',
            'construct_plasmid_aa_seq': 'MPKKKPTPIQLNPAPDGSAVNGTSSAETNLEALQKKLEELELDEQQRKRLEAFLTQKQKVGELKDDDFEKISELGAGNGGVVFKVSHKPSGLVMARKLIHLEIKPAIRNQIIRELQVLHECNSPYIVGFYGAFYSDGEISICMEHMDGGSLDQVLKKAGRIPEQILGKVSIAVIKGLTYLREKHKIMHRDVKPSNILVNSRGEIKLCDFGVSGQLIDSMANSFVGTRSYMSPERLQGTHYSVQSDIWSMGLSLVEMAVGRYPIPPPDAKELELMFGCQVEGDAAETPPRPRTPGRPLSSYGMDSRPPMAIFELLDYIVNEPPPKLPSGVFSLEFQDFVNKCLIKNPAERADLKQLMVHAFIKRSDAEEVDFAGWLCSTIGLNQPSTPTHAAGV',
        },
        {
            'targetid': 'KSYK_HUMAN_D0',
            'plasmid_source': 'addgene',
            'cloneID': '23907',
            'construct_plasmid_aa_seq': 'LDRKLLTLEDKELGSGNFGTVKKGYYQMKKVVKTVAVKILKNEANDPALKDELLAEANVMQQLDNPYIVRMIGICEAESWMLVMEMAELGPLNKYLQQNRHVKDKNIIELVHQVSMGMKYLEESNFVHRDLAARNVLLVTQHYAKISDFGLSKALRADENYYKAQTHGKWPVKWYAPECINYYKFSSKSDVWSFGVLMWEAFSYGQKPYRGMKGSEVTAMLEKGERMGCPAGCPREMYDLMNLCWTYDVENRPGFAAVELRLRNYYYDVVN',
        },
    ]

    addgene_targetids = [x.get('targetid') for x in targets if x.get('plasmid_source') == 'addgene']
    hip_targetids = [x.get('targetid') for x in targets if x.get('plasmid_source') == 'HIP']

    ws.cell(row=0, column=0).value = 'target ID'
    ws.cell(row=0, column=1).value = 'kinase family'
    ws.cell(row=0, column=2).value = 'plasmid source'
    ws.cell(row=0, column=3).value = 'plasmid ID'
    ws.cell(row=0, column=4).value = 'plate ID'
    ws.cell(row=0, column=5).value = 'well position'
    ws.cell(row=0, column=6).value = 'phosphatase to coexpress'
    ws.cell(row=0, column=7).value = 'aa start'
    ws.cell(row=0, column=8).value = 'aa end'
    ws.cell(row=0, column=9).value = 'dna start'
    ws.cell(row=0, column=10).value = 'dna end'
    ws.cell(row=0, column=11).value = 'construct aa seq'
    ws.cell(row=0, column=12).value = 'construct dna seq'
    ws.cell(row=0, column=13).value = 'original plasmid insert dna seq'

    for n,target in enumerate(targets):
        targetid = target.get('targetid')
        plasmid_source = target.get('plasmid_source')
        plasmid_ID = target.get('cloneID')
        construct_plasmid_aa_seq = target.get('construct_plasmid_aa_seq')

        print 'Working on target %s' % targetid

        if plasmid_source == 'addgene':
            plasmid_row = addgene_plasmids_df[addgene_plasmids_df.matching_targetID == targetid]
        elif plasmid_source == 'HIP':
            plasmid_row = hip_plasmids_df[hip_plasmids_df.matching_targetID == targetid]
        family = plasmid_row.UniProt_family.values[0]
        plateID = plasmid_row.plateID.values[0]
        well_pos = plasmid_row.well_pos.values[0]
        if family == 'TK':
            phosphatase_to_coexpress = 'YopH'
        else:
            phosphatase_to_coexpress = 'Lambda'
        plasmid_aa_seq = plasmid_row.construct_aa_seq.values[0]
        plasmid_dna_seq = plasmid_row.construct_dna_seq.values[0]
        plasmid_dna_orf_seq = plasmid_row.construct_dna_orf_seq.values[0]

        construct_start_plasmid_aa_coords = re.search(construct_plasmid_aa_seq, plasmid_aa_seq).start()
        construct_end_plasmid_aa_coords = re.search(construct_plasmid_aa_seq, plasmid_aa_seq).end() - 1

        construct_start_plasmid_dna_coords = construct_start_plasmid_aa_coords * 3
        construct_end_plasmid_dna_coords = construct_end_plasmid_aa_coords * 3 + 2
        construct_plasmid_dna_seq = plasmid_dna_orf_seq[construct_start_plasmid_dna_coords: construct_end_plasmid_dna_coords + 1]

        ws.cell(row=n+1, column=0).value = targetid
        ws.cell(row=n+1, column=1).value = family
        ws.cell(row=n+1, column=2).value = plasmid_source
        ws.cell(row=n+1, column=3).value = plasmid_ID
        ws.cell(row=n+1, column=4).value = plateID
        ws.cell(row=n+1, column=5).value = well_pos
        ws.cell(row=n+1, column=6).value = phosphatase_to_coexpress
        ws.cell(row=n+1, column=7).value = construct_start_plasmid_aa_coords + 1
        ws.cell(row=n+1, column=8).value = construct_end_plasmid_aa_coords + 1
        ws.cell(row=n+1, column=9).value = construct_start_plasmid_dna_coords + 1
        ws.cell(row=n+1, column=10).value = construct_end_plasmid_dna_coords + 1
        ws.cell(row=n+1, column=11).value = construct_plasmid_aa_seq
        ws.cell(row=n+1, column=12).value = construct_plasmid_dna_seq
        ws.cell(row=n+1, column=13).value = plasmid_dna_seq   # not just the ORF - this sequence will sometimes include stop codons and downstream sequence

        # import ipdb; ipdb.set_trace()


    wb.save(output_xl_filepath)

if __name__ == '__main__':
    main()
