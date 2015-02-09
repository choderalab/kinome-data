import sys, os, re, yaml
import numpy as np
from openpyxl import Workbook
import pandas as pd
import TargetExplorer
import Bio.Seq

# input params

input_filepath = 'selected-kinases.csv'
output_xl_filepath = 'selected-kinases.xlsx'
output_aln_filepath = 'selected-kinases.fa'

wb = Workbook()
ws = wb.get_active_sheet()
ws.title = 'kinase constructs'

df = pd.read_csv(input_filepath)

with open('manual_overrides-mk_spreadsheet.yaml', 'r') as manual_overrides_file:
    manual_overrides = yaml.load(manual_overrides_file)

print '%d kinase constructs found' % len(df)

# defs

def find_construct_seq_start(construct_seq, include_nterm_m=False):
    '''Searches for the start of a construct within a plasmid or PDB sequence
    which has been pre-aligned against the WT sequence. The aim is to avoid
    including expression tags.

    Returns the index denoting the start of the construct.

    Algorithm details:
    The preceding alignment marks conflicting residues in lower-case.
    This function iterates a sliding window over the sequence, checks for two
    matching residues at the beginning of the window, and then checks that
    there are less than a given maximum of conflicting residues within the
    window.

    include_nterm_m flag: includes single 'M' at beginning of SGC Oxford
    plasmids selected as constructs. This may be important for expression if no
    additional expression tags are to be added to the construct N-terminus.

    Example from SGC Oxford CLK3A-c005 aligned against CLK3_HUMAN
        ---mhHhHHhssgvdlgtenlyfqSmQSSKRSSRSVEDDKEGHLVCRIGDW
               x                  ^
        Construct start is defined at "^", not "x".

    Example from SGC Oxford CLK1A-c007 aligned against CLK1_HUMAN, with
    include_nterm_m=True
        ---mHLICQSGDVLSARYEIV
           ^
    '''
    window_size = 10
    lowercase_max_target = 3
    regex = re.compile('^[A-Z][A-Z]')
    sgc_regex = re.compile('^-m[A-Z]')

    if include_nterm_m:
        for w in range(0, len(construct_seq) - window_size):
            window = construct_seq[w: w + window_size]
            if w == 0:
                if re.match('^m[A-Z]', window):
                    num_lowercase = len([1 for char in window if not char.istitle()])
                    if num_lowercase < lowercase_max_target:
                        print 'NOTE: short N-terminal tag starting with "M" found:'
                        print construct_seq
                        return w
            else:
                if re.match(sgc_regex, window) != None:
                    num_lowercase = len([1 for char in window if not char.istitle()])
                    if num_lowercase < lowercase_max_target + 1:
                        print 'NOTE: short N-terminal tag starting with "M" found:'
                        print construct_seq
                        return w

    for w in range(0, len(construct_seq) - window_size):
        window = construct_seq[w: w + window_size]
        if re.match(regex, window) != None:
            num_lowercase = len([1 for char in window if not char.istitle()])
            if num_lowercase < lowercase_max_target:
                return w

    return None

# main

ws.cell(row=0, column=0).value = 'target ID'
ws.cell(row=0, column=1).value = 'kinase family'
ws.cell(row=0, column=2).value = 'plasmid source'
ws.cell(row=0, column=3).value = 'plasmid ID'
ws.cell(row=0, column=4).value = 'plate ID'
ws.cell(row=0, column=5).value = 'well position'
ws.cell(row=0, column=6).value = 'phosphatase to coexpress'
ws.cell(row=0, column=7).value = 'aa start'
ws.cell(row=0, column=8).value = 'aa end'
ws.cell(row=0, column=9).value = 'dna start'
ws.cell(row=0, column=10).value = 'dna end'
ws.cell(row=0, column=11).value = 'construct aa seq'
ws.cell(row=0, column=12).value = 'construct dna seq'
ws.cell(row=0, column=13).value = 'original plasmid insert dna seq'

with open(output_aln_filepath, 'w') as otxt_file:

    for n in range(len(df)):
        targetID = df['targetID'][n]
        print targetID
        family = df['family'][n] if df['family'][n] is not np.nan else ''
        if family == 'TK':
            phosphatase_to_coexpress = 'YopH'
        else:
            phosphatase_to_coexpress = 'Lambda'
        plasmid_source = df['plasmid_source'][n]
        plasmid_ID = df['plasmid_ID'][n]
        plateID = df['plateID'][n] if type(df['plateID'][n]) == str else ''
        well_pos = df['well_pos'][n] if type(df['well_pos'][n]) == str else ''
        construct_source = df['selected_construct_source'][n]

        # Modify with manual overrides
        if targetID in manual_overrides:
            for key, value in manual_overrides[targetID].iteritems():
                if key in ['plasmid_source', 'plasmid_ID', 'construct_source', 'plateID', 'well_pos']:
                    print 'Manual override: Setting %s to %s' % (key, value)
                    vars()[key] = value

        target_uniprot_seq = df['target_UniProt_seq'][n]
        plasmid_aa_seq = df['plasmid_aa_seq'][n].upper()
        plasmid_dna_orf_seq = df['plasmid_dna_orf_seq'][n].upper()
        plasmid_dna_seq = df['plasmid_dna_seq'][n].upper()
        pdb_aa_seq = df['top_pdb_aa_seq'][n]

        # Modify with manual overrides
        if targetID in manual_overrides:
            for key, value in manual_overrides[targetID].iteritems():
                if key in ['plasmid_dna_orf_seq', 'plasmid_dna_seq', 'plasmid_aa_seq']:
                    print 'Manual override: Setting %s to %s' % (key, value)
                    vars()[key] = value

        # Make sure aa seq is string
        try:
            assert type(plasmid_aa_seq) == str
        except:
            print targetID
            print type(plasmid_aa_seq)
            print df[ df['targetID'] == targetID ]
            raise

        # Align UniProt and plasmid sequences
        aln_ids = ['UniProt', 'plasmid']
        pre_aln_seqs = [target_uniprot_seq, plasmid_aa_seq]
        uniprot_plasmid_aln = TargetExplorer.align.run_clustalo(aln_ids, pre_aln_seqs) # list of aligned sequence strings

        # Align UniProt, plasmid and PDB sequences
        aln_ids = ['UniProt', 'plasmid', 'PDB']
        pre_aln_seqs = [target_uniprot_seq, plasmid_aa_seq, pdb_aa_seq]
        uniprot_plasmid_pdb_aln = TargetExplorer.align.run_clustalo(aln_ids, pre_aln_seqs)

        # make final-row aas lower-case if they conflict with first-row seq
        for i in range(len(uniprot_plasmid_aln[0])):
            uniprot_plasmid_aln[1] = list(uniprot_plasmid_aln[1])
            if uniprot_plasmid_aln[1][i] != uniprot_plasmid_aln[0][i]:
                uniprot_plasmid_aln[1][i] = uniprot_plasmid_aln[1][i].lower()
            uniprot_plasmid_aln[1] = ''.join(uniprot_plasmid_aln[1])
        for i in range(len(uniprot_plasmid_pdb_aln[0])):
            uniprot_plasmid_pdb_aln[2] = list(uniprot_plasmid_pdb_aln[2])
            if uniprot_plasmid_pdb_aln[2][i] != uniprot_plasmid_pdb_aln[0][i]:
                uniprot_plasmid_pdb_aln[2][i] = uniprot_plasmid_pdb_aln[2][i].lower()
            uniprot_plasmid_pdb_aln[2] = ''.join(uniprot_plasmid_pdb_aln[2])


        # If construct source is PDB, find start and end of construct in UniProt/plasmid/PDB alignment, removing conflicting sequence at termini
        if construct_source == 'PDB':
            construct_start_aln_coords = find_construct_seq_start(uniprot_plasmid_pdb_aln[2])
            construct_end_aln_coords = len(uniprot_plasmid_pdb_aln[2]) - find_construct_seq_start(uniprot_plasmid_pdb_aln[2][::-1]) - 1

            construct_plasmid_aa_seq_aln = uniprot_plasmid_pdb_aln[1][construct_start_aln_coords : construct_end_aln_coords+1]
            construct_plasmid_aa_seq = construct_plasmid_aa_seq_aln.replace('-', '')

        # If construct source is SGC, find start and end of construct in UniProt/plasmid alignment, removing conflicting sequence at termini
        elif construct_source == 'SGC':
            construct_start_aln_coords = find_construct_seq_start(uniprot_plasmid_aln[1], include_nterm_m=False)   # note: using N-terminal expression tag, so don't need to worry about N-terminal Met here
            construct_end_aln_coords = len(uniprot_plasmid_aln[1]) - find_construct_seq_start(uniprot_plasmid_aln[1][::-1]) - 1
            nterm_m_tag = re.search('m-*[-a-z]{8}-*[A-Z]{2}', uniprot_plasmid_aln[1])

            construct_plasmid_aa_seq_aln = uniprot_plasmid_aln[1][construct_start_aln_coords : construct_end_aln_coords+1]
            construct_plasmid_aa_seq = construct_plasmid_aa_seq_aln.replace('-', '').upper()


        construct_start_plasmid_aa_coords = re.search(construct_plasmid_aa_seq, plasmid_aa_seq).start()
        construct_end_plasmid_aa_coords = re.search(construct_plasmid_aa_seq, plasmid_aa_seq).end() - 1
        # print len(plasmid_aa_seq)*3, len(plasmid_dna_orf_seq)

        construct_start_plasmid_dna_coords = construct_start_plasmid_aa_coords * 3
        construct_end_plasmid_dna_coords = construct_end_plasmid_aa_coords * 3 + 2
        construct_plasmid_dna_seq = plasmid_dna_orf_seq[construct_start_plasmid_dna_coords: construct_end_plasmid_dna_coords + 1]

        if len(plasmid_aa_seq) * 3 != len(plasmid_dna_orf_seq):
            print 'WARNING for %s. len(aa) (%d) != len(dna)*3 (%d)' % (targetID, len(plasmid_aa_seq)*3, len(plasmid_dna_orf_seq))
            if plasmid_ID in ['HsCD00038083']:
                # Just need to remove one nucleotide from the end of these plasmid ORFs
                plasmid_dna_orf_seq = plasmid_dna_orf_seq[:-1]
                print 'Ok. Checked manually - continuing...'
            else:
                print plasmid_ID
                print Bio.Seq.Seq(plasmid_dna_orf_seq, Bio.Alphabet.generic_dna).translate()
                print plasmid_aa_seq
                print plasmid_dna_orf_seq
                raise Exception

        ws.cell(row=n+1, column=0).value = targetID
        ws.cell(row=n+1, column=1).value = family
        ws.cell(row=n+1, column=2).value = plasmid_source
        ws.cell(row=n+1, column=3).value = plasmid_ID
        ws.cell(row=n+1, column=4).value = plateID
        ws.cell(row=n+1, column=5).value = well_pos
        ws.cell(row=n+1, column=6).value = phosphatase_to_coexpress
        ws.cell(row=n+1, column=7).value = construct_start_plasmid_aa_coords + 1
        ws.cell(row=n+1, column=8).value = construct_end_plasmid_aa_coords + 1
        ws.cell(row=n+1, column=9).value = construct_start_plasmid_dna_coords + 1
        ws.cell(row=n+1, column=10).value = construct_end_plasmid_dna_coords + 1
        ws.cell(row=n+1, column=11).value = construct_plasmid_aa_seq
        ws.cell(row=n+1, column=12).value = construct_plasmid_dna_seq
        ws.cell(row=n+1, column=13).value = plasmid_dna_seq   # not just the ORF - this sequence will sometimes include stop codons and downstream sequence

        result_string = '>%s  %s  %s  %s\n' % (targetID, construct_source, plasmid_source, plasmid_ID)
        if construct_source == 'PDB':
            result_string += '%s\n%s\n%s\n\n' % (uniprot_plasmid_pdb_aln[0], uniprot_plasmid_pdb_aln[1], uniprot_plasmid_pdb_aln[2])
        if construct_source == 'SGC':
            result_string += '%s\n%s\n\n' % (uniprot_plasmid_aln[0], uniprot_plasmid_aln[1])

        otxt_file.write(result_string)

wb.save(output_xl_filepath)
