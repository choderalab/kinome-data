import os, re
import argparse
import openpyxl
import pandas as pd
from lxml import etree
import Bio.Seq

# ========
# Command-line args and parameters
# ========

argparser = argparse.ArgumentParser()
argparser.add_argument('--database_path', type=str, help='Path to a TargetExplorer database XML file', required=True)
args = argparser.parse_args()

expression_system_regex = 'E\. coli'

desired_taxid = 9606

# ========
# Read in plasmid spreadsheet
# ========

plasmid_lib_filepath = os.path.join('oxford sgc clones for sbs_020513.xlsx')
wb = openpyxl.load_workbook(plasmid_lib_filepath)
ws = wb.get_sheet_by_name(name='Sheet1')
nrows = ws.get_highest_row()

# ========
# DataFrame for plasmid data
# ========

plasmid_df = {
'cloneID':[],
'HGNCSymbol':[],
'UniProtAC':[],
'UniProt_entry_name':[],
'UniProt_family':[],
'DB_target_rank':[],
'construct_dna_seq':[],
'construct_dna_orf_seq':[],
'construct_aa_seq':[],
'construct_aa_seq_w_tags':[],
'Protein family':[],
'Expression system':[],
'Expression cell line':[],
'Availability comments':[],
'Special expression comments':[],
'Vector name':[],
'Antibiotic resistance':[],
'Expression tag':[],
'Vector comments':[],
'Mutations/sequence comments':[],
}

# ========
# Output columns for text file
# ========

txt_output_cols = ['HGNCSymbol', 'UniProtAC', 'UniProt_entry_name', 'construct_aa_seq', 'Protein family', 'DB_target_rank']

# ========
# Read in database
# ========

DB_root = etree.parse(args.database_path)

# ========
# Iterate through plasmid spreadsheet
# ========

for row in range(2, nrows-1):
    cloneID = ws.cell('C%d' % row).value
    if cloneID == None:
        continue
    expression_system = ws.cell('J%d' % row).value
    if not re.match(expression_system_regex, expression_system):
        continue
    HGNCSymbol = ws.cell('G%d' % row).value

    # Search for Hugo Symbol in database
    matching_DB_entry = DB_root.find('entry/UniProt[@NCBI_taxID="%d"]/../HGNC/entry[@Approved_Symbol="%s"]/../..' % (desired_taxid, HGNCSymbol))
    if matching_DB_entry == None:
        continue

    UniProtAC = matching_DB_entry.find('UniProt').get('AC')
    UniProt_entry_name = matching_DB_entry.find('UniProt').get('entry_name')
    UniProt_family = matching_DB_entry.find('UniProt').get('family')
    target_rank = matching_DB_entry.find('target_score/domain')
    if target_rank != None:
        target_rank = target_rank.get('target_rank')

    construct_aa_seq_w_tags = ws.cell('T%d' % row).value
    construct_dna_seq = ws.cell('U%d' % row).value
    # get DNA ORF (assume starts at index 0)
    construct_dna_orf_seq = construct_dna_seq
    for n in range(0, len(construct_dna_seq), 3):
        if construct_dna_seq[n:n+3] in ['TAG', 'TAA', 'TGA']:
            construct_dna_orf_seq = construct_dna_seq[:n]
            break

    construct_aa_seq = Bio.Seq.Seq(construct_dna_orf_seq, Bio.Alphabet.generic_dna).translate()

    # if UniProt_entry_name == 'MRCKA_HUMAN':
    #     print construct_aa_seq
    #     print Bio.Seq.Seq(construct_dna_seq, Bio.Alphabet.generic_dna).translate()

    n_tag = ws.cell('P%d' % row).value
    c_tag = ws.cell('Q%d' % row).value

    plasmid_df['cloneID'].append(cloneID)
    plasmid_df['HGNCSymbol'].append(HGNCSymbol)
    plasmid_df['UniProtAC'].append(UniProtAC)
    plasmid_df['UniProt_entry_name'].append(UniProt_entry_name)
    plasmid_df['UniProt_family'].append(UniProt_family)
    plasmid_df['DB_target_rank'].append(target_rank)
    plasmid_df['construct_aa_seq_w_tags'].append( construct_aa_seq_w_tags )
    plasmid_df['construct_aa_seq'].append( construct_aa_seq )
    plasmid_df['construct_dna_seq'].append( construct_dna_seq )
    plasmid_df['construct_dna_orf_seq'].append( construct_dna_orf_seq )
    plasmid_df['Protein family'].append( ws.cell('E%d' % row).value )
    plasmid_df['Expression system'].append( expression_system )
    plasmid_df['Expression cell line'].append( ws.cell('K%d' % row).value )
    plasmid_df['Availability comments'].append( ws.cell('L%d' % row).value )
    plasmid_df['Special expression comments'].append( ws.cell('M%d' % row).value )
    plasmid_df['Vector name'].append( ws.cell('N%d' % row).value )
    plasmid_df['Antibiotic resistance'].append( ws.cell('O%d' % row).value )
    plasmid_df['Vector comments'].append( ws.cell('R%d' % row).value )
    plasmid_df['Mutations/sequence comments'].append( ws.cell('S%d' % row).value )
    assert not (n_tag != None and c_tag != None)
    if n_tag != None:
        plasmid_df['Expression tag'].append( 'N-term %s' % n_tag )
    elif c_tag != None:
        plasmid_df['Expression tag'].append( 'C-term %s' % c_tag )
    else:
        plasmid_df['Expression tag'].append( None )

plasmid_df = pd.DataFrame(plasmid_df)
plasmid_df.set_index('cloneID', inplace=True)

plasmid_df.to_csv('plasmid-data.csv')
with open('plasmid-data.txt', 'w') as plasmid_data_txt_file:
    plasmid_data_txt_file.write(plasmid_df.to_string(columns=txt_output_cols))
