import os, re, argparse
import openpyxl
from openpyxl import Workbook
import Bio, Bio.Seq, Bio.Alphabet
from lxml import etree
from lxml.builder import E
import TargetExplorer as clab
import pandas as pd

plasmid_library_dir = '.'
Harvard_plasmid_library_filepath = os.path.join(plasmid_library_dir, 'Mehle_Kinase_VS1_pJP1520_new_plates.xlsx')

# ========
# Command-line args
# ========

argparser = argparse.ArgumentParser()
argparser.add_argument('--database_path', type=str, help='Path to a TargetExplorer database XML file', required=True)
args = argparser.parse_args()

desired_taxid = 9606
ignore_clones = ['HsCD00038286', 'HsCD00037967'] # both of these have deletions in the kinase catalytic domain

# ==========
# Parse plasmid library spreadsheet
# ==========

wb = openpyxl.load_workbook(Harvard_plasmid_library_filepath)
sheet_ranges = wb.get_sheet_by_name(name = 'Kinase_VS1_pJP1520_new_plates')
nrows = sheet_ranges.get_highest_row()
#plasmidIDs = {}
#plasmid_aa_seqs = {}
#plasmid_dna_seqs = {}
plasmid_df = {
'cloneID':[],
'dna_seq':[],
'aa_seq':[],
'NCBI_GeneID':[],
'Symbol':[]
}
for row in range(2,nrows):
    cloneID = sheet_ranges.cell('E%d' % row).value    # type(int)
    if cloneID in ignore_clones:
        continue
    NCBI_GeneID = sheet_ranges.cell('H%d' % row).value    # type(int)
    Symbol = sheet_ranges.cell('I%d' % row).value
    dna_seq = sheet_ranges.cell('L%d' % row).value 
    if len(dna_seq) % 3 != 0:
        print 'WARNING: length of DNA sequence not divisible by 3, for plasmid with Gene Symbol %s and NCBI Gene ID %s' % (Symbol, NCBI_GeneID)
    # Translate to DNA sequence (don't include stop codon)
    aa_seq = Bio.Seq.Seq(dna_seq, Bio.Alphabet.generic_dna).translate(to_stop=True)
    #plasmidIDs[NCBI_GeneID] = plasmidID
    #plasmid_aa_seqs[NCBI_GeneID] = aa_seq
    #plasmid_dna_seqs[NCBI_GeneID] = dna_seq
    plasmid_df['cloneID'].append(cloneID)
    plasmid_df['dna_seq'].append(dna_seq)
    plasmid_df['aa_seq'].append(aa_seq)
    plasmid_df['NCBI_GeneID'].append(NCBI_GeneID)
    plasmid_df['Symbol'].append(Symbol)

#plasmid_NCBI_GeneIDs = plasmid_aa_seqs.keys()

plasmid_df = pd.DataFrame(plasmid_df)

DB_root = etree.parse(args.database_path).getroot()


# To be used to construct a pandas DataFrame
data_fields = ['cloneID', 'NCBI_GeneID', 'orig_gene_symbol', 'UniProtAC', 'UniProt_entry_name', 'UniProt_family', 'insert_dna_seq', 'insert_aa_seq']
output_data = pd.DataFrame( [['None'] * len(data_fields)] * len(plasmid_df), columns=data_fields)

#DB_gene_name_nodes = [ gene_name_node for gene_name_node in DB_root.findall('entry/UniProt/gene_names/gene_name') ]

# ===========
# Iterate through plasmids
# ===========

for p in plasmid_df.index:
    cloneID = plasmid_df['cloneID'][p]
    insert_dna_seq = plasmid_df['dna_seq'][p]
    insert_aa_seq = plasmid_df['aa_seq'][p]
    plasmid_NCBI_GeneID = plasmid_df['NCBI_GeneID'][p]
    plasmid_Symbol = plasmid_df['Symbol'][p]

    output_data['cloneID'][p] = cloneID
    output_data['NCBI_GeneID'][p] = plasmid_NCBI_GeneID
    output_data['orig_gene_symbol'][p] = plasmid_Symbol
    output_data['insert_dna_seq'][p] = insert_dna_seq
    output_data['insert_aa_seq'][p] = insert_aa_seq


    # find matching DB entry via NCBI GeneID
    DB_entry = DB_root.find('entry/UniProt[@NCBI_taxID="%d"]/../NCBI_Gene/entry[@ID="%s"]/../..' % (desired_taxid, plasmid_NCBI_GeneID))
    if DB_entry == None:
        print 'Matching DB entry not found for Gene ID %s cloneID %s Symbol %s' % (plasmid_NCBI_GeneID, cloneID, plasmid_Symbol)

        # check if the plasmid gene symbol matches in any in the DB
        matching_DB_gene_names = DB_root.xpath('entry/UniProt/gene_names/gene_name[text()="%s"]' % plasmid_Symbol)
        if len(matching_DB_gene_names) > 0:
            for matching_DB_gene_name in matching_DB_gene_names:
                matching_DB_UniProt_entry_name = matching_DB_gene_name.getparent().getparent().get('entry_name')
                print 'NOTE: plasmid gene symbol %s for clone %s appears to match a gene symbol in the DB entry %s' % (plasmid_Symbol, cloneID, matching_DB_UniProt_entry_name)

        continue


    DB_UniProt_node = DB_entry.find('UniProt')
    DB_domains = DB_UniProt_node.findall('domains/domain[@targetID]')
    UniProtAC = DB_UniProt_node.get('AC')
    UniProt_entry_name = DB_UniProt_node.get('entry_name')
    UniProt_family = DB_UniProt_node.get('family')
    UniProt_canonseq = clab.core.sequnwrap( DB_UniProt_node.find('isoforms/canonical_isoform/sequence').text )

    output_data['UniProtAC'][p] = UniProtAC
    output_data['UniProt_entry_name'][p] = UniProt_entry_name
    output_data['UniProt_family'][p] = UniProt_family

    #break

# construct pandas DataFrame and write to csv

output_data.set_index('cloneID', inplace=True)
output_data.to_csv('plasmid-data.csv')

